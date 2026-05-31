from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from datetime import datetime
import os


class ReportGenerator:
    """Generate Excel reports from analysis results"""

    def __init__(self, reports_dir: str = "./reports"):
        self.reports_dir = reports_dir
        os.makedirs(reports_dir, exist_ok=True)

    def generate_excel_report(self, analysis: Dict[str, Any], analysis_id: str) -> str:
        """
        Generate comprehensive Excel report with all metrics
        Returns path to generated file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "10-K Analysis"

        # Define all columns as specified
        columns = [
            ("Ticker", 10),
            ("Company", 25),
            ("Year", 8),
            ("Sector", 20),

            # Segment & Concentration
            ("Segment Mix", 30),
            ("Geo Concentration", 25),
            ("Top Customer %", 12),

            # Revenue Metrics
            ("Revenue ($M)", 12),
            ("Organic Growth %", 15),
            ("FX Impact ($M)", 12),
            ("M&A Impact ($M)", 12),

            # Backlog/RPO
            ("Backlog/RPO ($M)", 15),
            ("Backlog YoY %", 12),

            # Margins
            ("Gross Margin %", 12),
            ("Operating Margin %", 15),

            # Cash Flow
            ("FCF ($M)", 12),
            ("FCF Margin %", 12),
            ("FCF/NI Ratio", 12),

            # SBC
            ("SBC ($M)", 10),
            ("SBC % Revenue", 12),
            ("Dilution %", 10),

            # CapEx
            ("CapEx ($M)", 12),

            # Working Capital
            ("DSO (days)", 10),
            ("DIO (days)", 10),
            ("DPO (days)", 10),
            ("CCC (days)", 10),

            # Debt & Leverage
            ("Net Debt ($M)", 12),
            ("Debt/EBITDA", 12),
            ("Interest Coverage", 15),
            ("<24m Maturities ($M)", 18),

            # Capital Allocation
            ("Buybacks ($M)", 12),
            ("Dividends ($M)", 12),
            ("Shareholder Yield %", 18),

            # Deferred Revenue
            ("Def Revenue ($M)", 15),
            ("Def Revenue YoY %", 15),

            # Quality Metrics
            ("Accruals Ratio", 12),

            # Off-Balance
            ("Off-Balance Commit ($M)", 20),

            # Risk Indicators
            ("Material Weakness", 18),
            ("New Top Risks", 12),
            ("Concentration Risk", 15),

            # Sector Specific
            ("RPO ($M)", 12),
            ("Current RPO ($M)", 15),
            ("Rule of 40", 12),
            ("Book-to-Bill", 12),

            # Catalysts
            ("Catalysts", 50),

            # Scoring
            ("Catalyst Score /40", 18),
            ("Quality Score /35", 16),
            ("Risk Score /25", 14),
            ("Total Score /100", 15),
            ("Rating", 12),

            # Red Flags
            ("Red Flags", 60),
        ]

        # Write header row
        for col_idx, (col_name, col_width) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # Extract data from analysis
        metrics = analysis.get('financial_metrics', {})
        sector_metrics = analysis.get('sector_metrics', {})
        risks = analysis.get('risk_indicators', {})
        trade_score = analysis.get('trade_score', {})

        # Helper to format percentage
        def fmt_pct(val):
            if val is None:
                return "N/A"
            return f"{val * 100:.1f}%"

        # Helper to format number
        def fmt_num(val, decimals=1):
            if val is None:
                return "N/A"
            return f"{val:.{decimals}f}"

        # Helper to format currency
        def fmt_curr(val, decimals=0):
            if val is None:
                return "N/A"
            return f"${val:,.{decimals}f}M"

        # Prepare segment mix string
        segments = analysis.get('segments', {})
        segment_str = ", ".join([f"{k}: {v.get('revenue', 0):.0f}M ({fmt_pct(v.get('growth', 0))})"
                                  for k, v in segments.items()]) if segments else "N/A"

        # Prepare geo concentration string
        geo_conc = analysis.get('geographic_concentration', {})
        geo_str = ", ".join([f"{k}: {v:.1%}" for k, v in geo_conc.items()]) if geo_conc else "N/A"

        # Top customer percentage
        top_customers = analysis.get('top_customers', [])
        top_cust_str = max([c.get('revenue_pct', 0) for c in top_customers], default=0) if top_customers else 0

        # Catalysts string
        catalysts = analysis.get('catalysts', [])
        catalyst_str = "; ".join(catalysts[:5]) if catalysts else "None identified"

        # Red flags string
        red_flags = risks.get('red_flags', [])
        red_flag_str = "; ".join(red_flags) if red_flags else "None"

        # RPO or Backlog
        rpo_backlog = sector_metrics.get('rpo') or sector_metrics.get('backlog')
        rpo_backlog_yoy = sector_metrics.get('backlog_yoy', 0)

        # Write data row
        row_data = [
            analysis.get('ticker', 'N/A'),
            analysis.get('company_name', 'N/A'),
            analysis.get('fiscal_year', 'N/A'),
            analysis.get('sector', 'N/A'),

            # Segment & Concentration
            segment_str,
            geo_str,
            fmt_pct(top_cust_str) if isinstance(top_cust_str, (int, float)) else top_cust_str,

            # Revenue
            fmt_curr(metrics.get('revenue')),
            fmt_pct(metrics.get('organic_growth')),
            fmt_curr(metrics.get('fx_impact')),
            fmt_curr(metrics.get('ma_impact')),

            # Backlog/RPO
            fmt_curr(rpo_backlog),
            fmt_pct(rpo_backlog_yoy),

            # Margins
            fmt_pct(metrics.get('gross_margin')),
            fmt_pct(metrics.get('operating_margin')),

            # Cash Flow
            fmt_curr(metrics.get('fcf')),
            fmt_pct(metrics.get('fcf_margin')),
            fmt_num(metrics.get('fcf_to_ni'), 2),

            # SBC
            fmt_curr(metrics.get('sbc_amount')),
            fmt_pct(metrics.get('sbc_pct_revenue')),
            fmt_pct(metrics.get('dilution_pct')),

            # CapEx
            fmt_curr(metrics.get('capex')),

            # Working Capital
            fmt_num(metrics.get('dso'), 0),
            fmt_num(metrics.get('dio'), 0),
            fmt_num(metrics.get('dpo'), 0),
            fmt_num(metrics.get('ccc'), 0),

            # Debt
            fmt_curr(metrics.get('net_debt')),
            fmt_num(metrics.get('debt_to_ebitda'), 1),
            fmt_num(metrics.get('interest_coverage'), 1),
            fmt_curr(risks.get('debt_maturity_24m')),

            # Capital Allocation
            fmt_curr(metrics.get('buybacks')),
            fmt_curr(metrics.get('dividends')),
            fmt_pct(metrics.get('shareholder_yield')),

            # Deferred Revenue
            fmt_curr(metrics.get('deferred_revenue')),
            fmt_pct(metrics.get('deferred_revenue_yoy')),

            # Quality
            fmt_num(metrics.get('accruals_ratio'), 3),

            # Off-Balance
            fmt_curr(risks.get('off_balance_commitments')),

            # Risk Indicators
            "Yes" if risks.get('material_weakness') else "No",
            "Yes" if risks.get('new_top_risks') else "No",
            "Yes" if risks.get('customer_concentration') else "No",

            # Sector Specific
            fmt_curr(sector_metrics.get('rpo')),
            fmt_curr(sector_metrics.get('current_rpo')),
            fmt_num(sector_metrics.get('rule_of_40'), 1),
            fmt_num(sector_metrics.get('book_to_bill'), 2),

            # Catalysts
            catalyst_str,

            # Scoring
            fmt_num(trade_score.get('catalyst_trend_score', 0), 1),
            fmt_num(trade_score.get('quality_cash_score', 0), 1),
            fmt_num(trade_score.get('risk_score', 0), 1),
            fmt_num(trade_score.get('total_score', 0), 1),
            trade_score.get('rating', 'N/A'),

            # Red Flags
            red_flag_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Add color coding for rating
        rating_col = len(columns) - 1  # Rating is second to last
        rating_cell = ws.cell(row=2, column=rating_col)
        rating = trade_score.get('rating', 'N/A')

        if rating == "Strong Buy":
            rating_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            rating_cell.font = Font(bold=True, color="FFFFFF")
        elif rating == "Buy":
            rating_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif rating == "Hold":
            rating_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        elif rating == "Avoid":
            rating_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            rating_cell.font = Font(bold=True, color="FFFFFF")

        # Create insights sheet
        insights_ws = wb.create_sheet("AI Insights")
        insights_ws.column_dimensions['A'].width = 120

        insights_ws['A1'] = "AI Analysis Insights"
        insights_ws['A1'].font = Font(bold=True, size=14)

        insights_ws['A3'] = analysis.get('ai_insights', 'No insights available')
        insights_ws['A3'].alignment = Alignment(wrap_text=True, vertical="top")

        insights_ws['A5'] = "Key Takeaways:"
        insights_ws['A5'].font = Font(bold=True, size=12)

        takeaways = analysis.get('key_takeaways', [])
        for idx, takeaway in enumerate(takeaways, start=6):
            insights_ws[f'A{idx}'] = f"• {takeaway}"
            insights_ws[f'A{idx}'].alignment = Alignment(wrap_text=True, vertical="top")

        # Save file
        filename = f"10k_analysis_{analysis_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)

        return filepath
